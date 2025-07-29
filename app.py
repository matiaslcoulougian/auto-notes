import streamlit as st
import pandas as pd
import yfinance as yf
import requests
from bs4 import BeautifulSoup
import time
from openpyxl.styles import PatternFill
import io
import datetime

# Configure Streamlit page to wide mode
st.set_page_config(
    page_title="Structured Investment Pro",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# --- Common headers for web scraping ---
BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "Accept-Language": "en-US,en;q=0.9,es;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
    "Sec-Ch-Ua": '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
    "Sec-Ch-Ua-Mobile": "?0",
    "Sec-Ch-Ua-Platform": '"Windows"',
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
    "Referer": "https://www.google.com/",
    "Connection": "keep-alive",
}


# --- Función auxiliar para obtener el mínimo de 52 semanas desde Yahoo Finance ---
def obtener_min_52_semanas(ticker):
    """
    Scrapes Yahoo Finance to get the 52-week low value from the 52 Week Range field.
    Returns the lower value from a range like "177.00 - 488.54" -> 177.00
    """
    url = f"https://finance.yahoo.com/quote/{ticker}"
    
    try:
        response = requests.get(url, headers=BROWSER_HEADERS, timeout=10)
        soup = BeautifulSoup(response.text, "html.parser")
        
        # Look for the 52 Week Range element
        # Find the fin-streamer with data-field="fiftyTwoWeekRange"
        range_element = soup.find("fin-streamer", {"data-field": "fiftyTwoWeekRange"})
        
        if range_element:
            range_text = range_element.get_text(strip=True)
            # Parse the range like "177.00 - 488.54"
            if " - " in range_text:
                parts = range_text.split(" - ")
                if len(parts) == 2:
                    try:
                        min_52_week = float(parts[0].strip())
                        return round(min_52_week, 2)
                    except ValueError:
                        pass
        
        # Fallback: look for any element with title "52 Week Range"
        range_label = soup.find("span", {"title": "52 Week Range"})
        if range_label:
            # Find the value span that follows
            parent = range_label.find_parent()
            if parent:
                value_span = parent.find("span", class_="value")
                if value_span:
                    fin_streamer = value_span.find("fin-streamer")
                    if fin_streamer:
                        range_text = fin_streamer.get_text(strip=True)
                        if " - " in range_text:
                            parts = range_text.split(" - ")
                            if len(parts) == 2:
                                try:
                                    min_52_week = float(parts[0].strip())
                                    return round(min_52_week, 2)
                                except ValueError:
                                    pass
        
        return None
        
    except Exception as e:
        print(f"Error scraping 52-week low para {ticker}: {e}")
        return None


# --- Función para buscar datos en Yahoo Finance ---
def obtener_datos_yahoo(ticker):
    try:
        data = yf.Ticker(ticker)
        hist = data.history(period="1y")
        if hist.empty:
            return None, None, None, None

        precio_actual = round(hist['Close'].iloc[-1], 2)
        hace_1_anio = round(hist['Close'].iloc[0], 2)
        # Get 52-week low from Yahoo Finance scraping
        min_1y = obtener_min_52_semanas(ticker)
        min_1y = round(min_1y, 2) if min_1y is not None else None
        target_yhoo = data.info.get('targetMeanPrice', None)
        target_yhoo = round(target_yhoo, 2) if target_yhoo is not None else None

        return precio_actual, target_yhoo, hace_1_anio, min_1y
    except Exception as e:
        print(f"Error en Yahoo para {ticker}: {e}")
        return None, None, None, None


# --- Scraping Target Price Morgan Stanley ---
def obtener_target_morgan(ticker):
    url = f"https://www.tipranks.com/stocks/{ticker.lower()}/forecast"
    try:
        response = requests.get(url, headers=BROWSER_HEADERS, timeout=10)
        soup = BeautifulSoup(response.text, "html.parser")
        
        # Find the table body with analyst data (React Table structure)
        table_body = soup.find("div", class_="rt-tbody")
        if table_body:
            # Find all table rows (rt-tr-group contains each analyst row)
            row_groups = table_body.find_all("div", class_="rt-tr-group")
            
            for row_group in row_groups:
                # Find the actual row within the group
                row = row_group.find("div", class_="rt-tr")
                if row:
                    # Find all table cells
                    cells = row.find_all("div", class_="rt-td")
                    
                    if len(cells) >= 3:  # Make sure we have enough columns
                        # Expert Firm is the second column (index 1)
                        expert_firm_cell = cells[1]
                        expert_firm = expert_firm_cell.get_text(strip=True)
                        
                        # Check if this row is for Morgan Stanley
                        if "Morgan Stanley" in expert_firm:
                            # Price Target is the third column (index 2)
                            price_target_cell = cells[2]
                            
                            # Look for the price target value within the cell
                            # Handle both single values and ranges (e.g., $80 → $85)
                            price_spans = price_target_cell.find_all("span", class_="Mdcvgxd7")
                            
                            if price_spans:
                                try:
                                    if len(price_spans) == 1:
                                        # Single price target
                                        price_target_text = price_spans[0].get_text(strip=True)
                                        price_target = float(price_target_text.replace("$", "").replace(",", ""))
                                        return round(price_target, 2)
                                    else:
                                        # Range: take the higher value (second span) for conservative estimate
                                        high_target_text = price_spans[1].get_text(strip=True)
                                        price_target = float(high_target_text.replace("$", "").replace(",", ""))
                                        return round(price_target, 2)
                                except ValueError:
                                    pass
                            else:
                                # Fallback: get all text from the cell and parse ranges
                                price_target_text = price_target_cell.get_text(strip=True)
                                if price_target_text and price_target_text != "—":
                                    try:
                                        # Check if it's a range with arrow or dash
                                        if "→" in price_target_text:
                                            # Split by arrow and take the higher value
                                            parts = price_target_text.split("→")
                                            if len(parts) == 2:
                                                high_value = parts[1].strip().replace("$", "").replace(",", "")
                                                price_target = float(high_value)
                                                return round(price_target, 2)
                                        elif "–" in price_target_text or "-" in price_target_text:
                                            # Handle dash ranges like $80-$85
                                            separator = "–" if "–" in price_target_text else "-"
                                            parts = price_target_text.split(separator)
                                            if len(parts) == 2:
                                                high_value = parts[1].strip().replace("$", "").replace(",", "")
                                                price_target = float(high_value)
                                                return round(price_target, 2)
                                        else:
                                            # Single value
                                            price_target = float(price_target_text.replace("$", "").replace(",", ""))
                                            return round(price_target, 2)
                                    except ValueError:
                                        pass
        return None
    except Exception as e:
        print(f"Error scraping Morgan Stanley para {ticker}: {e}")
        return None


# -- Cálculo de Score (según lógica del PRD) --
def calcular_score(nota, pesos):
    # Extraer valores, convertir None a 0
    tasa = nota.get("Tasa") or 0
    colchon = nota.get("Colchón") or 0
    memory = 1 if nota.get("Memory") else 0
    precio_actual = nota.get("Precio actual") or 0
    target_yahoo = nota.get("Target Yahoo") or 0
    target_morgan = nota.get("Target MS") or 0
    hace_1_anio = nota.get("Hace 1 año") or 0
    min_1_anio = nota.get("Mín 1 año") or 0

    # Pesos
    p_tasa = pesos["Tasa"]
    p_colchon = pesos["Colchón"]
    p_memory = pesos["Memory"]
    p_yahoo = pesos["Target Yahoo"]
    p_ms = pesos["Target MS"]
    p_1y = pesos["1 Año"]
    p_min1y = pesos["Mín 1 Año"]

    # Trigger
    try:
        trigger = precio_actual * (100 - colchon) / 100 if precio_actual and colchon is not None else 0
    except Exception:
        trigger = 0

    # Evitar divisiones por cero
    def safe_div(n, d):
        try:
            return n / d if d else 0
        except:
            return 0

    # Términos polinómicos
    t1 = tasa * p_tasa / 20
    t2 = colchon * p_colchon / 100
    t3 = memory * p_memory
    t4 = ((safe_div(target_yahoo, precio_actual) - 1) * p_yahoo if precio_actual else 0)
    t5 = ((safe_div(target_morgan, precio_actual) - 1) * p_ms if precio_actual else 0)
    t6 = (safe_div(hace_1_anio, trigger) * p_1y if trigger else 0)
    t7 = (safe_div(min_1_anio, trigger) * p_min1y if trigger else 0)

    score = t1 + t2 + t3 + t4 + t5 + t6 + t7
    return round(score, 2)

# Título de la app
st.title("Structured Investment Pro 📈")

# -- Session State para persistencia de datos temporales --
if "notas" not in st.session_state:
    st.session_state["notas"] = []
if "pesos" not in st.session_state:
    st.session_state["pesos"] = {
        "Tasa": 0.15,
        "Colchón": 0.34,
        "Memory": 0.24,
        "Target Yahoo": 0.09,
        "Target MS": 0.09,
        "1 Año": 0.09,
        "Mín 1 Año": 0.09,
    }
if "reset" not in st.session_state:
    st.session_state["reset"] = False
if "edit_mode" not in st.session_state:
    st.session_state["edit_mode"] = False

# -- Formulario de Inputs --
with st.form("input_form", clear_on_submit=True):
    st.subheader("Agregar nueva nota")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        ticker = st.text_input("Ticker", max_chars=8)
    with col2:  # format in 0.00
        tasa = st.number_input("Tasa (%)", min_value=0.0, max_value=100.0, step=1.0, format="%.2f")
    with col3:
        colchon = st.number_input("Colchón (%)", min_value=0.0, max_value=100.0, step=1.0, format="%.2f")
    with col4:
        memory = st.checkbox("Memory", value=False)
    submitted = st.form_submit_button("Agregar nota")

    # Al agregar, guardamos en session_state y buscamos datos automáticamente
    if submitted:
        if ticker and len(st.session_state["notas"]) < 20:
            # Mostrar spinner mientras se buscan los datos
            with st.spinner(f"Agregando {ticker.upper()}..."):
                # Crear nota básica
                nueva_nota = {
                    "Ticker": ticker.upper(),
                    "Tasa": tasa,
                    "Colchón": colchon,
                    "Memory": memory,
                }
                
                # Buscar datos de internet automáticamente
                try:
                    # --- Yahoo Finance ---
                    precio_actual, target_yhoo, hace_1_anio, min_1y = obtener_datos_yahoo(ticker.upper())
                    # --- Morgan Stanley ---
                    target_morgan = obtener_target_morgan(ticker.upper())
                    
                    # Agregar datos obtenidos a la nota
                    nueva_nota["Precio actual"] = precio_actual
                    nueva_nota["Target Yahoo"] = target_yhoo
                    nueva_nota["Hace 1 año"] = hace_1_anio
                    nueva_nota["Mín 1 año"] = min_1y
                    nueva_nota["Target MS"] = target_morgan
                    
                    # Calcular score automáticamente
                    score = calcular_score(nueva_nota, st.session_state["pesos"])
                    nueva_nota["Score"] = score
                    
                    st.session_state["notas"].append(nueva_nota)
                    st.success(f"✅ {ticker.upper()} listo con score {score}!")
                    
                except Exception as e:
                    # Si hay error en la búsqueda, agregar solo los datos básicos con score
                    # Ensure all fields exist even if they're None
                    nueva_nota["Precio actual"] = None
                    nueva_nota["Target Yahoo"] = None
                    nueva_nota["Hace 1 año"] = None
                    nueva_nota["Mín 1 año"] = None
                    nueva_nota["Target MS"] = None
                    
                    score = calcular_score(nueva_nota, st.session_state["pesos"])
                    nueva_nota["Score"] = score
                    st.session_state["notas"].append(nueva_nota)
                    st.warning(f"⚠️ {ticker.upper()} agregado con score {score}, pero hubo problemas obteniendo algunos datos de internet.")
                    
        elif not ticker:
            st.warning("Ingrese un ticker válido.")
        elif len(st.session_state["notas"]) >= 20:
            st.warning("Máximo 20 notas.")

# -- Edición de Pesos Ponderados --
with st.expander("⚙️ Configurar pesos del motor"):
    st.write("Modifique los pesos de cada variable (suma no obligatoria = 1):")
    for key in st.session_state["pesos"]:
        valor = st.number_input(
            f"Peso para {key}",
            min_value=0.0, max_value=1.0, step=0.01,
            value=float(st.session_state["pesos"][key]),
            key=f"peso_{key}"
        )
        st.session_state["pesos"][key] = valor

# -- Visualización de la tabla de notas --
st.subheader("Notas cargadas")
if st.session_state["notas"]:
    df = pd.DataFrame(st.session_state["notas"])
    
    # Check if we have internet-sourced data to edit
    has_internet_data = any(
        any(key in nota for key in ["Precio actual", "Target Yahoo", "Hace 1 año", "Mín 1 año", "Target MS"])
        for nota in st.session_state["notas"]
    )
    
    # Funciones para el semáforo
    def calcular_percentil_score(val, scores):
        """
        Calcula el percentil relativo de un valor dentro del rango de scores.
        Retorna un valor entre 0 y 1.
        """
        if len(scores) == 0:
            return 1
        
        max_score = max(scores)
        min_score = min(scores)
        
        if max_score - min_score > 0:
            return (val - min_score) / (max_score - min_score)
        else:
            return 1  # Todos los scores son iguales, asignar el mejor color

    def color_semaforo(val):
        # Colores tipo semáforo usando percentiles relativos (igual que Excel)
        if "Score" in df.columns:
            try:
                scores = df["Score"].dropna().values
                if len(scores) > 0:
                    percentil = calcular_percentil_score(val, scores)
                    
                    if percentil >= 0.66:
                        return "background-color: #99FF99; color: black;"  # Verde (mismo que Excel)
                    elif percentil >= 0.33:
                        return "background-color: #FFFF99; color: black;"  # Amarillo (mismo que Excel)
                    else:
                        return "background-color: #FF9999; color: black;"  # Rojo (mismo que Excel)
            except:
                return ""
        return ""
    
    # Mostrar tabla única con semáforo si hay scores
    if "Score" in df.columns:
        # Formatear columnas numéricas a 2 decimales
        numeric_columns = ["Tasa", "Colchón", "Precio actual", "Target Yahoo", "Hace 1 año", "Mín 1 año", "Target MS", "Score"]
        format_dict = {}
        for col in numeric_columns:
            if col in df.columns:
                # Use lambda to handle None values properly
                format_dict[col] = lambda x: f"{x:.2f}" if pd.notna(x) else ""
        
        st.dataframe(
            df.style.map(color_semaforo, subset=["Score"]).format(format_dict),
            use_container_width=True
        )
    else:
        # Formatear columnas numéricas a 2 decimales (sin semáforo)
        numeric_columns = ["Tasa", "Colchón", "Precio actual", "Target Yahoo", "Hace 1 año", "Mín 1 año", "Target MS", "Score"]
        format_dict = {}
        for col in numeric_columns:
            if col in df.columns:
                # Use lambda to handle None values properly
                format_dict[col] = lambda x: f"{x:.2f}" if pd.notna(x) else ""
        
        st.dataframe(df.style.format(format_dict), use_container_width=True)
    
    # Botones debajo de la tabla
    col1, col2, col3 = st.columns([2, 1, 1])
    
    # Show edit button if there's internet data
    if has_internet_data:
        with col1:
            if st.button("✏️ Editar", help="Editar valores obtenidos de internet", key="edit_btn"):
                st.session_state["edit_mode"] = True
                st.rerun()
    
    # Botón para exportar a Excel (solo si hay scores)
    if "Score" in df.columns and not df["Score"].isnull().all():
        def exportar_excel_semaforo(df):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Notas")
                workbook = writer.book
                worksheet = writer.sheets["Notas"]
                # Determinar la columna "Score"
                score_col = None
                for idx, col in enumerate(df.columns, 1):
                    if col == "Score":
                        score_col = idx
                        break
                # Solo si hay Score y más de 1 nota
                if score_col is not None and len(df) > 0:
                    scores = df["Score"].dropna().values
                    for row in range(2, len(df) + 2):  # Desde fila 2 (1 es header)
                        cell = worksheet.cell(row=row, column=score_col)
                        val = cell.value
                        if val is not None:
                            # Usar la misma lógica de percentiles que la UI
                            percentil = calcular_percentil_score(val, scores)
                            
                            if percentil >= 0.66:
                                fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")  # Verde
                            elif percentil >= 0.33:
                                fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Amarillo
                            else:
                                fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Rojo
                            cell.fill = fill
            output.seek(0)
            return output
        
        with col2:
            excel_data = exportar_excel_semaforo(df)
            today = datetime.datetime.now().strftime("%Y-%m-%d")
            st.download_button(
                label="📥 Descargar Excel",
                data=excel_data,
                file_name="notas_scoring_{}.xlsx".format(today),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    st.info(f"Total notas cargadas: {len(st.session_state['notas'])}/20")
    
else:
    st.info("No hay notas cargadas aún.")

# -- Modal de edición (fuera del bloque condicional) --
# Show modal if edit mode is active
if st.session_state.get("edit_mode", False):
    # -- Modal de edición para datos obtenidos de internet --
    @st.dialog("✏️ Editar datos obtenidos de internet")
    def edit_modal():
        st.info("Modifica solo los valores que necesites corregir. Los campos vacíos mantendrán su valor original.")
        
        # Manejo de eliminación fuera del formulario
        confirm_delete_index = st.session_state.get("confirm_delete_index", None)
        if confirm_delete_index is not None:
            ticker_to_delete = st.session_state["notas"][confirm_delete_index]["Ticker"]
            st.warning(f"⚠️ ¿Estás seguro de eliminar {ticker_to_delete}?")
            col1, col2, col3 = st.columns([1, 1, 2])
            with col1:
                if st.button("✅", key="confirm_delete_modal"):
                    st.session_state["notas"].pop(confirm_delete_index)
                    st.session_state["confirm_delete_index"] = None
                    # Si no quedan notas, cerrar el modal
                    if len(st.session_state["notas"]) == 0:
                        st.session_state["edit_mode"] = False
                        st.success(f"🗑️ {ticker_to_delete} eliminado. Modal cerrado (no quedan notas).")
                    else:
                        st.success(f"🗑️ {ticker_to_delete} eliminado correctamente!")
                    st.rerun()
            with col2:
                if st.button("❌", key="cancel_delete_modal"):
                    st.session_state["confirm_delete_index"] = None
                    st.rerun()
        
        # Botones de eliminación individuales
        if len(st.session_state["notas"]) > 0:
            st.subheader("Eliminar notas")
            cols = st.columns(len(st.session_state["notas"]))
            for i, nota in enumerate(st.session_state["notas"]):
                with cols[i]:
                    if st.button(f"🗑️ {nota['Ticker']}", key=f"delete_modal_{i}", help=f"Eliminar {nota['Ticker']}"):
                        st.session_state["confirm_delete_index"] = i
                        st.rerun()
        else:
            st.info("No hay notas para eliminar.")
        
        st.divider()
        
        # Create edit form
        with st.form("edit_form"):
            edited_notas = []
            
            for i, nota in enumerate(st.session_state["notas"]):
                st.write(f"**{nota['Ticker']}**")
                
                # Primera fila: Campos originales editables
                col1, col2, col3 = st.columns(3)
                with col1:
                    tasa_edit = st.number_input(
                        "Tasa (%)",
                        value=float(nota.get("Tasa", 0)),
                        min_value=0.0,
                        max_value=100.0,
                        step=1.0,
                        format="%.2f",
                        key=f"tasa_{i}",
                        help="Tasa de interés"
                    )
                with col2:
                    colchon_edit = st.number_input(
                        "Colchón (%)",
                        value=float(nota.get("Colchón", 0)),
                        min_value=0.0,
                        max_value=100.0,
                        step=1.0,
                        format="%.2f",
                        key=f"colchon_{i}",
                        help="Porcentaje de colchón"
                    )
                with col3:
                    memory_edit = st.checkbox(
                        "Memory",
                        value=nota.get("Memory", False),
                        key=f"memory_{i}",
                        help="Indicador de memoria"
                    )
                
                # Segunda fila: Campos de datos de internet
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    precio_actual = st.number_input(
                        "Precio actual",
                        value=float(nota.get("Precio actual", 0)) if nota.get("Precio actual") else 0.0,
                        step=0.01,
                        format="%.2f",
                        key=f"precio_{i}",
                        help="Precio actual de la acción"
                    )
                with col2:
                    target_yahoo = st.number_input(
                        "Target Yahoo",
                        value=float(nota.get("Target Yahoo", 0)) if nota.get("Target Yahoo") else 0.0,
                        step=0.01,
                        format="%.2f",
                        key=f"yahoo_{i}",
                        help="Precio objetivo de Yahoo Finance"
                    )
                with col3:
                    hace_1_anio = st.number_input(
                        "Hace 1 año",
                        value=float(nota.get("Hace 1 año", 0)) if nota.get("Hace 1 año") else 0.0,
                        step=0.01,
                        format="%.2f",
                        key=f"anio_{i}",
                        help="Precio hace 1 año"
                    )
                with col4:
                    min_1_anio = st.number_input(
                        "Mín 1 año",
                        value=float(nota.get("Mín 1 año", 0)) if nota.get("Mín 1 año") else 0.0,
                        step=0.01,
                        format="%.2f",
                        key=f"min_{i}",
                        help="Mínimo en 1 año (52-week low)"
                    )
                with col5:
                    target_morgan = st.number_input(
                        "Target MS",
                        value=float(nota.get("Target MS", 0)) if nota.get("Target MS") else 0.0,
                        step=0.01,
                        format="%.2f",
                        key=f"morgan_{i}",
                        help="Precio objetivo de Morgan Stanley"
                    )
                
                # Create updated note
                nota_editada = nota.copy()
                nota_editada["Tasa"] = tasa_edit
                nota_editada["Colchón"] = colchon_edit
                nota_editada["Memory"] = memory_edit
                nota_editada["Precio actual"] = precio_actual if precio_actual > 0 else None
                nota_editada["Target Yahoo"] = target_yahoo if target_yahoo > 0 else None
                nota_editada["Hace 1 año"] = hace_1_anio if hace_1_anio > 0 else None
                nota_editada["Mín 1 año"] = min_1_anio if min_1_anio > 0 else None
                nota_editada["Target MS"] = target_morgan if target_morgan > 0 else None
                edited_notas.append(nota_editada)
                if i < len(st.session_state["notas"]) - 1:
                    st.divider()
            
            # Form buttons
            col1, col2, col3 = st.columns([1, 1, 2])
            with col1:
                if st.form_submit_button("Guardar", type="primary"):
                    # Recalcular scores para todas las notas editadas
                    for nota in edited_notas:
                        score = calcular_score(nota, st.session_state["pesos"])
                        nota["Score"] = score
                    
                    st.session_state["notas"] = edited_notas
                    st.session_state["edit_mode"] = False
                    st.success("Cambios guardados y scores recalculados exitosamente!")
                    st.rerun()
            with col2:
                if st.form_submit_button("Cancelar"):
                    st.session_state["edit_mode"] = False
                    st.rerun()
    edit_modal()



# -- Firma del autor --
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: #666; font-style: italic; margin-top: 2rem;'>"
    "Creado por Mati Coulougian. Con amor para Leo ♥️📈"
    "</div>", 
    unsafe_allow_html=True
)
